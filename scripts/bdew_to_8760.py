# scripts/bdew_xlsx_to_8760.py
# -*- coding: utf-8 -*-
"""
BDEW-SLP Excel -> 8760 JSON (extrem robust)
- unterstützt vertikale Blöcke, breite Tabellen & Block-Matrix
- füllt merged cells auf
- Stunden 0..23 / 1..24 / "00:00".. "23:00"
- Tagtypen WT/SA/SO/FT (FT wird wie SO behandelt)
- Feiertage je Bundesland mit 'holidays'
Abhängigkeiten:
    pip install openpyxl holidays
"""
import argparse, json, math
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from holidays import Germany

GER_MONTHS = {
    "januar":1, "februar":2, "märz":3, "maerz":3, "april":4,
    "mai":5, "juni":6, "juli":7, "august":8, "september":9,
    "oktober":10, "november":11, "dezember":12
}
ENG_MONTHS = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12
}
DAYTYPES = {
    # Werktag
    "WT":"WT","WKT":"WT","WkT":"WT","WKt":"WT","Werktag":"WT","Weekday":"WT",
    # Samstag
    "SA":"SA","Samstag":"SA","Saturday":"SA",
    # Sonntag/Feiertag
    "SO":"SO","Sonntag":"SO","So":"SO","SUN":"SO","Sunday":"SO",
    "FT":"SO","Feiertag":"SO","Holiday":"SO"
}

def detect_month(v):
    if v is None: return None
    s = str(v).strip().lower().replace(".","")
    return GER_MONTHS.get(s) or ENG_MONTHS.get(s)

def detect_daytype(v):
    if v is None: return None
    s = str(v).strip()
    return DAYTYPES.get(s)

def to_float(x):
    if x is None: return math.nan
    if isinstance(x,(int,float)): return float(x)
    try:
        return float(str(x).replace(",","."))
    except:
        return math.nan

def is_hour_label(x):
    if x is None: return False
    s = str(x).strip().lower()
    # 0..23
    if s.isdigit():
        try:
            h = int(s); return 0 <= h <= 24
        except: return False
    # 00:00..23:00
    if s.endswith(":00"):
        try:
            h = int(s.split(":",1)[0]); return 0 <= h <= 24
        except: return False
    return False

def hour_value(x):
    """Mappe "0", "00:00", "1", "24" -> int Stunde (0..23) oder None."""
    if x is None: return None
    s = str(x).strip().lower()
    try:
        if s.endswith(":00"): s = s.split(":",1)[0]
        h = int(s)
        if 0 <= h <= 23: return h
        if h == 24: return 0  # 24 = 0 Uhr (manchmal in Tabellen)
    except: pass
    return None

def fill_merged(ws):
    """kopiert Werte aus merged cell top-left in alle Zellen des Merge-Bereichs"""
    grid = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            for r in range(1, ws.max_row+1)]
    for rng in ws.merged_cells.ranges:
        minr, minc, maxr, maxc = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        val = ws.cell(row=minr, column=minc).value
        for r in range(minr, maxr+1):
            for c in range(minc, maxc+1):
                grid[r-1][c-1] = val
    return grid  # 0-indexiert [r-1][c-1]

# ---------- Variante A: vertikale Blöcke ----------
def parse_vertical(grid):
    m = {}
    R, C = len(grid), (len(grid[0]) if grid else 0)
    for r in range(0, R-26):
        for c in range(0, C):
            mon = detect_month(grid[r][c])
            if not mon: 
                continue
            dtp = detect_daytype(grid[r+1][c])
            if not dtp:
                continue
            vals = [to_float(grid[r+2+i][c]) for i in range(24)]
            if sum(1 for v in vals if math.isfinite(v)) >= 20:
                # Lücken füllen
                for i,v in enumerate(vals):
                    if not math.isfinite(v):
                        prev = next((vals[j] for j in range(i-1,-1,-1) if math.isfinite(vals[j])), None)
                        nxt  = next((vals[j] for j in range(i+1,24) if math.isfinite(vals[j])), None)
                        if prev is not None and nxt is not None: vals[i] = (prev+nxt)/2
                        elif prev is not None: vals[i] = prev
                        elif nxt  is not None: vals[i] = nxt
                        else: vals[i] = 0.0
                m[(mon, dtp)] = vals
    return m

# ---------- Variante B/C: breite Tabelle / Block-Matrix ----------
def parse_wide(grid):
    """
    Sucht Zeile mit vielen Monatstiteln, darunter (1-3 Zeilen tiefer) Tagtyp-Köpfe,
    links Stunden-Spalte (0..23 / 1..24 / 00:00..23:00).
    Liest für jede gefundene (Monat,Tagtyp)-Spalte 24 Werte.
    """
    m = {}
    R, C = len(grid), (len(grid[0]) if grid else 0)
    if R == 0 or C == 0: return m

    # 1) Monatstitel-Zeile finden (>=6 Treffer)
    header_r = None
    months_at_col = {}
    for r in range(min(60,R)):
        found = []
        for c in range(C):
            mon = detect_month(grid[r][c])
            if mon: found.append((c, mon))
        if len(found) >= 6:
            header_r = r
            months_at_col = {c:mon for c,mon in found}
            break
    if header_r is None: 
        return m

    # 2) Tagtyp-Zeile: header_r+1 .. header_r+3 durchsuchen
    typ_r = None
    for rr in range(header_r+1, min(header_r+4, R)):
        # wenigstens 2 Tagtypen finden
        if sum(1 for c in range(C) if detect_daytype(grid[rr][c])) >= 2:
            typ_r = rr; break
    if typ_r is None: 
        return m

    # 3) Stunden-Spalte links suchen (viele hour labels)
    hour_c = None
    best_hits = -1
    for c in range(C):
        hits = 0
        for rr in range(typ_r+1, min(typ_r+1+40, R)):
            if is_hour_label(grid[rr][c]): hits += 1
        if hits > best_hits:
            best_hits, hour_c = hits, c
    if hour_c is None or best_hits < 10:
        return m

    # 4) Für jede Spalte, die Monatstitel oben hat, suche Tagtyp in typ_r +/- 2 Spalten
    for c0, mon in months_at_col.items():
        # In einigen Tabellen stehen Tagtypen 0..3 Spalten rechts von der Monatsspalte
        chosen = []
        for c in range(c0, min(c0+6, C)):
            dtp = detect_daytype(grid[typ_r][c])
            if dtp:
                chosen.append((c, dtp))
        # fallback: schaue auch 3 Spalten links (manchmal versetzt)
        if not chosen:
            for c in range(max(0,c0-3), c0):
                dtp = detect_daytype(grid[typ_r][c])
                if dtp:
                    chosen.append((c, dtp))
        # falls wir gar nichts finden, weiter
        if not chosen: 
            continue

        # 5) Startzeile der Stunden: erste Zeile mit "0/1/00:00" in hour_c unterhalb typ_r
        start_r = None
        for rr in range(typ_r+1, min(typ_r+60, R)):
            hv = hour_value(grid[rr][hour_c])
            if hv in (0,1):  # viele Tabellen starten mit 0 oder 1
                start_r = rr; break
        if start_r is None:
            # fallback: erste Zeile, die überhaupt ein Stundenlabel hat
            for rr in range(typ_r+1, min(typ_r+60, R)):
                if is_hour_label(grid[rr][hour_c]): 
                    start_r = rr; break
        if start_r is None: 
            continue

        # 6) jetzt je (Spalte, Tagtyp) die 24 Werte abgreifen
        for c, dtp in chosen:
            vals = []
            for k in range(24):
                rr = start_r + k
                if rr >= R: vals.append(math.nan); continue
                vals.append(to_float(grid[rr][c]))
            # Lücken füllen
            for i,v in enumerate(vals):
                if not math.isfinite(v):
                    prev = next((vals[j] for j in range(i-1,-1,-1) if math.isfinite(vals[j])), None)
                    nxt  = next((vals[j] for j in range(i+1,24) if math.isfinite(vals[j])), None)
                    if prev is not None and nxt is not None: vals[i] = (prev+nxt)/2
                    elif prev is not None: vals[i] = prev
                    elif nxt  is not None: vals[i] = nxt
                    else: vals[i] = 0.0
            m[(mon, dtp)] = vals

    return m

def build_8760(lut, year:int, bundesland:str):
    de_holidays = Germany(years=year, prov=bundesland)
    hours = []
    dt = datetime(year,1,1,0,0)
    end= datetime(year+1,1,1,0,0)
    while dt < end:
        mon = dt.month
        wd  = dt.weekday()  # 0=Mo … 6=So
        dtp = "SO" if (wd==6 or dt.date() in de_holidays) else ("SA" if wd==5 else "WT")
        block = lut.get((mon, dtp)) or lut.get((mon,"WT")) or [0.0]*24
        hours.append(float(block[dt.hour]))
        dt += timedelta(hours=1)
    return hours

def normalize(values, mode:str):
    total = sum(values)
    if total <= 0: raise ValueError("Summe der Werte ist 0 – Excel prüfen.")
    if mode == "share":
        return [v/total for v in values], "share_of_year"
    if mode == "annual_1e6":
        factor = 1_000_000/total
        return [v*factor for v in values], "kWh_per_1e6"
    return values, "raw"

def main():
    ap = argparse.ArgumentParser(description="BDEW SLP Excel -> 8760 JSON (robust)")
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--year",  type=int, required=True)
    ap.add_argument("--bundesland", required=True)  # z. B. RP, NW, BY
    ap.add_argument("--out",  required=True)
    ap.add_argument("--normalize", choices=["share","annual_1e6","none"], default="share")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{args.sheet}' nicht gefunden. Verfügbar: {wb.sheetnames}")
    ws  = wb[args.sheet]

    grid = fill_merged(ws)           # merged cells auffüllen
    lut  = parse_vertical(grid)
    if not lut:
        lut = parse_wide(grid)
    if not lut:
        raise ValueError("Konnte keine Monats/Tagtyp-Blöcke erkennen (weder vertikal noch breit). Bitte Excel prüfen.")

    hours = build_8760(lut, args.year, args.bundesland)
    hours, unit = normalize(hours, args.normalize)

    out = {
        "source":"BDEW-2025",
        "profile": args.sheet,
        "bundesland": args.bundesland,
        "year": args.year,
        "unit": unit,
        "hours": hours
    }
    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    with open(args.out,"w",encoding="utf-8") as f:
        json.dump(out,f,ensure_ascii=False)
    print(f"OK → {args.out} (len={len(hours)}, sum={sum(hours):.6f}, unit={unit})")

if __name__ == "__main__":
    main()
